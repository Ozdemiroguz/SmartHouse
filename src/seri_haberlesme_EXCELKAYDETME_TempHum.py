import serial
import time
import openpyxl


arduino = serial.Serial("COM4", baudrate=960, timeout=1)


wb=openpyxl.Workbook()

sheet=wb.active

sheet["A1"]="Saat"
sheet["B1"]="Sıcaklık"
sheet["C1"]="Nem"

satir=2

while satir<102:
    data = arduino.readline().decode('utf-8').strip().split("\t")
    if len(data) == 2:
        humidity, temperature = map(float, data)
        simdi = time.strftime("%H:%M:%S")
        print(f"Saat {simdi}: Sıcaklık={temperature} Nem={humidity}")

        sheet.cell(row=satir, column=1, value=simdi)
        sheet.cell(row=satir, column=2, value=temperature)
        sheet.cell(row=satir, column=3, value=humidity)

        satir += 1

    time.sleep(1)  


wb.save("Sicaklik_Nem_Zaman.xlsx")